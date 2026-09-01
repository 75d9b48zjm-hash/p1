"""Generator template pembukuan UMKM berbasis Excel (tanpa macro).

Menghasilkan di /app/excel_templates:
  - Pembukuan-Template.xlsx
  - Pembukuan-TokoMaju-Contoh.xlsx
  - Rekap-Admin.xlsx

Desain: bersih-profesional, kartu KPI + panah naik/turun vs bulan lalu,
grafik bawaan Excel + "grafik dalam sel" (REPT) yang tetap aman di HP
(Google Sheets / Excel Mobile). Tanpa VBA/macro.
"""
import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = "/app/excel_templates"
ROWS = 500          # baris data transaksi (2..501)
LAST = ROWS + 1

# ------------------------------------------------------------------ warna
NAVY = "0F2B46"     # judul & header tabel
INK = "1F2937"      # teks utama
SLATE = "64748B"    # teks sekunder
FAINT = "94A3B8"    # teks bantu
LINE = "E2E8F0"     # garis tabel
BG = "F8FAFC"       # latar lembut
WHITE = "FFFFFF"

GREEN = "047857"
GREEN_BG = "ECFDF5"
RED = "B91C1C"
RED_BG = "FEF2F2"
BLUE = "1D4ED8"
BLUE_BG = "EFF6FF"
AMBER = "B45309"
AMBER_BG = "FFFBEB"
TEAL = "0E7C7B"
GOLD = "C8961E"
CREAM = "FBF7EF"
YELLOW = "FFF3C4"

FONT = "Calibri"
RP = '"Rp"\\ #,##0'
NUM = '#,##0'
DATEF = 'dd/mm/yyyy'
BAR = "\u2588"      # blok penuh untuk grafik dalam sel

INCOME_CATS = ["Penjualan Tunai", "Penjualan Online", "Pendapatan Jasa",
               "Titipan/Konsinyasi", "Pendapatan Lain-lain"]
EXPENSE_CATS = ["Bahan Baku", "Barang Dagang", "Gaji & Upah", "Sewa Tempat",
                "Listrik & Air", "Transport & Kirim", "Kemasan", "Pulsa & Internet",
                "Peralatan", "Promosi", "Pajak & Retribusi", "Pengeluaran Lain-lain"]
METHODS = ["Tunai", "Transfer Bank", "QRIS", "E-wallet", "Lainnya"]
STATUSES = ["Belum Ditinjau", "Disetujui", "Perlu Perbaikan"]

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# alamat sel kendali di sheet Dashboard (dipakai sheet lain)
D_MONTH = "$C$6"
D_YEAR = "$F$6"
D_KEY = "$C$8"
D_KEY_PREV = "$F$8"
DG = "'Data Grafik'"

MONTHS_FULL = ('"Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus",'
               '"September","Oktober","November","Desember"')
MONTHS_SHORT = '"Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"'


# ------------------------------------------------------------------ helper
def no_grid(ws):
    ws.sheet_view.showGridLines = False


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def block(ws, r1, c1, r2, c2, bg=None, accent=None):
    """Kotak rapi: garis tipis di sekeliling + garis aksen tebal di atas."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = Border(
                left=thin if c == c1 else None,
                right=thin if c == c2 else None,
                top=(Side(style="thick", color=accent) if accent else thin) if r == r1 else None,
                bottom=thin if r == r2 else None,
            )


def page_head(ws, eyebrow, title, subtitle, col=2, span=5, title_formula=None):
    """Kepala halaman: label kecil, judul besar, keterangan."""
    c1 = get_column_letter(col)
    end = get_column_letter(col + span - 1)
    ws.row_dimensions[1].height = 8
    ws[f"{c1}2"] = eyebrow
    ws[f"{c1}2"].font = Font(name=FONT, size=8, bold=True, color=TEAL)
    ws[f"{c1}3"] = title_formula or title
    ws[f"{c1}3"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
    ws.row_dimensions[3].height = 30
    ws[f"{c1}4"] = subtitle
    ws[f"{c1}4"].font = Font(name=FONT, size=9, color=SLATE)
    for r in (2, 3, 4):
        ws.merge_cells(f"{c1}{r}:{end}{r}")
    ws.row_dimensions[5].height = 6


def section(ws, row, col, span, text):
    c1 = get_column_letter(col)
    end = get_column_letter(col + span - 1)
    ws.merge_cells(f"{c1}{row}:{end}{row}")
    cell = ws[f"{c1}{row}"]
    cell.value = text
    cell.font = Font(name=FONT, size=9, bold=True, color=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20
    for c in range(col, col + span):
        ws.cell(row=row, column=c).border = Border(bottom=Side(style="medium", color=NAVY))


def month_picker(ws, month, year, note):
    """Baris kendali: Bulan + Tahun + nama bulan + kunci bulan (baris 8 disembunyikan)."""
    ws["B6"] = "Bulan (1-12)"
    ws["E6"] = "Tahun"
    for a in ("B6", "E6"):
        ws[a].font = Font(name=FONT, size=9, bold=True, color=SLATE)
    for a, v in (("C6", month), ("F6", year)):
        cell = ws[a]
        cell.value = v
        cell.font = Font(name=FONT, size=13, bold=True, color=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=AMBER_BG)
        cell.border = Border(left=Side(style="thin", color=GOLD), right=Side(style="thin", color=GOLD),
                             top=Side(style="thin", color=GOLD), bottom=Side(style="thin", color=GOLD))
    ws.row_dimensions[6].height = 24
    ws.merge_cells("B7:C7")
    ws["B7"] = f'=CHOOSE({D_MONTH},{MONTHS_FULL})&" "&{D_YEAR}'
    ws["B7"].font = Font(name=FONT, size=12, bold=True, color=TEAL)
    ws.merge_cells("E7:F7")
    ws["E7"] = note
    ws["E7"].font = Font(name=FONT, size=9, italic=True, color=FAINT)
    ws.row_dimensions[7].height = 20
    ws["B8"] = "kunci"
    ws["C8"] = f'=TEXT(DATE({D_YEAR},{D_MONTH},1),"yyyy-mm")'
    ws["E8"] = "kunci bulan lalu"
    ws["F8"] = f'=TEXT(DATE({D_YEAR},{D_MONTH}-1,1),"yyyy-mm")'
    for a in ("B8", "C8", "E8", "F8"):
        ws[a].font = Font(name=FONT, size=8, color=FAINT)
    ws.row_dimensions[8].hidden = True
    ws.row_dimensions[9].height = 8


def card(ws, row, col, label, value, cur_ref=None, prev_ref=None, fmt=RP,
         accent=BLUE, bg=BLUE_BG, good_when_up=True, note=None):
    """Kartu KPI 3 baris x 2 kolom: label, angka besar, indikator vs bulan lalu."""
    c1, c2 = col, col + 1
    block(ws, row, c1, row + 2, c2, bg=bg, accent=accent)
    for r, h in ((row, 16), (row + 1, 30), (row + 2, 17)):
        ws.row_dimensions[r].height = h
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

    lab = ws.cell(row=row, column=c1, value=label)
    lab.font = Font(name=FONT, size=8, bold=True, color=accent)
    lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    val = ws.cell(row=row + 1, column=c1, value=value)
    val.font = Font(name=FONT, size=17, bold=True, color=INK)
    val.number_format = fmt
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    sub = ws.cell(row=row + 2, column=c1)
    sub.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    sub.font = Font(name=FONT, size=9, color=SLATE)
    if cur_ref and prev_ref:
        sub.value = (f'=IF({prev_ref}=0,"belum ada pembanding bulan lalu",'
                     f'IF({cur_ref}>={prev_ref},"\u25b2 ","\u25bc ")'
                     f'&TEXT(ABS({cur_ref}-{prev_ref})/ABS({prev_ref}),"0%")&" vs bulan lalu")')
        rng = f"{get_column_letter(c1)}{row + 2}:{get_column_letter(c2)}{row + 2}"
        up_color, down_color = (GREEN, RED) if good_when_up else (RED, GREEN)
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({prev_ref}<>0,{cur_ref}>={prev_ref})'],
            font=Font(name=FONT, size=9, bold=True, color=up_color), stopIfTrue=True))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({prev_ref}<>0,{cur_ref}<{prev_ref})'],
            font=Font(name=FONT, size=9, bold=True, color=down_color), stopIfTrue=True))
    elif note:
        sub.value = note


def mini(ws, row, col, label, formula, fmt=NUM, bg=BG):
    """Kotak statistik kecil: label di kiri, angka di kanan."""
    lab = ws.cell(row=row, column=col, value=label)
    lab.font = Font(name=FONT, size=9, color=SLATE)
    lab.alignment = Alignment(vertical="center", indent=1)
    val = ws.cell(row=row, column=col + 1, value=formula)
    val.font = Font(name=FONT, size=10, bold=True, color=INK)
    val.number_format = fmt
    val.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    block(ws, row, col, row, col + 1, bg=bg)
    ws.row_dimensions[row].height = 20


def table_head(ws, row, col, labels, fill=NAVY, height=26):
    for i, v in enumerate(labels):
        c = ws.cell(row=row, column=col + i, value=v)
        c.font = Font(name=FONT, bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height


def kv(ws, row, label, value=None, fmt=None, label_col="B", val_col="C"):
    ws[f"{label_col}{row}"] = label
    ws[f"{label_col}{row}"].font = Font(name=FONT, bold=True, size=10, color=NAVY)
    ws[f"{label_col}{row}"].alignment = Alignment(vertical="center")
    cell = ws[f"{val_col}{row}"]
    if value is not None:
        cell.value = value
    if fmt:
        cell.number_format = fmt
    cell.fill = PatternFill("solid", fgColor=CREAM)
    cell.border = BORDER
    cell.font = Font(name=FONT, size=10, color=INK)
    cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    return cell


def note_line(ws, row, col, span, text, color=FAINT):
    c1 = get_column_letter(col)
    end = get_column_letter(col + span - 1)
    ws.merge_cells(f"{c1}{row}:{end}{row}")
    ws[f"{c1}{row}"] = text
    ws[f"{c1}{row}"].font = Font(name=FONT, size=9, italic=True, color=color)
    ws[f"{c1}{row}"].alignment = Alignment(vertical="center", wrap_text=False)


def fit_print(ws, landscape=False):
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if landscape:
        ws.page_setup.orientation = "landscape"


# ---------------------------------------------------------------- sheets UMKM
def sheet_petunjuk(wb):
    ws = wb.create_sheet("Petunjuk")
    ws.sheet_properties.tabColor = NAVY
    no_grid(ws)
    widths(ws, {"A": 2, "B": 4, "C": 96})
    page_head(ws, "KASUMKM \u00b7 BUKU KAS UMKM", "Cara Pakai File Ini",
              "Cukup 4 kolom setiap kali mencatat. Sisanya dihitung otomatis.", col=2, span=2)

    groups = [
        ("UNTUK PEMILIK USAHA", TEAL, [
            "Buka sheet TRANSAKSI (lihat tab di bawah).",
            "Isi baris kosong paling bawah: Tanggal, Jenis (Masuk/Keluar), Nominal, Keterangan.",
            "Kolom Kategori, Metode Bayar, dan Status BOLEH DIKOSONGKAN \u2014 nanti diisi admin.",
            "Nominal ditulis angka saja, tanpa Rp dan tanpa titik. Contoh: 150000",
            "Punya foto nota? Simpan di Google Drive, tempel tautannya di kolom Tautan Nota.",
            "Catat setiap hari walau hanya satu transaksi. Tidak perlu rapi, yang penting tercatat.",
        ]),
        ("UNTUK ADMIN / PEMBUKU", BLUE, [
            "Baris berwarna kuning = kategori belum diisi. Isi lewat dropdown di kolom Kategori.",
            "Setelah benar, ubah Status menjadi 'Disetujui'. Kalau ragu pilih 'Perlu Perbaikan' "
            "dan tulis pertanyaan di kolom Catatan Admin.",
            "Sheet CEK DATA menampilkan baris bermasalah (nominal kosong, tanggal aneh, duplikat).",
            "Sheet DASHBOARD: ubah Bulan & Tahun \u2014 kartu, indikator, dan grafik ikut berubah.",
            "Akhir bulan: cetak sheet LABA RUGI menjadi PDF dan kirim ke pemilik usaha.",
        ]),
        ("CARA BACA DASHBOARD", GREEN, [
            "Empat kartu di atas = Uang Masuk, Uang Keluar, Laba/Rugi, Saldo Kas bulan terpilih.",
            "\u25b2 hijau = membaik dibanding bulan lalu, \u25bc merah = menurun. "
            "Untuk Uang Keluar warnanya dibalik (naik = merah).",
            "Lampu status: hijau SEHAT (laba positif), kuning IMBANG, merah WASPADA (boros).",
            "Balok \u2588 pada 'Uang Paling Banyak Keluar' menunjukkan besar pengeluaran per kategori.",
            "Grafik di sebelah kanan: 12 bulan, tren harian, dan komposisi pengeluaran.",
        ]),
        ("PENTING", AMBER, [
            "Jangan menghapus atau memindahkan kolom, karena rumus mengacu ke posisi kolom.",
            "Kolom J (Bulan) di sheet Transaksi dan sheet DATA GRAFIK adalah pembantu rumus \u2014 biarkan saja.",
            "Kapasitas 500 baris transaksi. Kalau penuh, buat file baru untuk tahun berikutnya.",
            "Sel rumus dikunci agar tidak tertimpa. Buka lewat Review \u2192 Unprotect Sheet (tanpa password).",
            "Ingin diisi dari HP? Unggah ke Google Drive, buka dengan Google Spreadsheet.",
        ]),
    ]
    r = 6
    for head, color, items in groups:
        section(ws, r, 2, 2, head)
        ws[f"B{r}"].font = Font(name=FONT, size=9, bold=True, color=color)
        r += 1
        for i, text in enumerate(items, start=1):
            ws.cell(row=r, column=2, value=f"{i}.").font = Font(name=FONT, size=9, bold=True, color=color)
            c = ws.cell(row=r, column=3, value=text)
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 26 if len(text) > 95 else 18
            r += 1
        r += 1
    return ws


def sheet_profil(wb, profil=None):
    ws = wb.create_sheet("Profil Usaha")
    ws.sheet_properties.tabColor = TEAL
    no_grid(ws)
    widths(ws, {"A": 2, "B": 34, "C": 42})
    page_head(ws, "KASUMKM \u00b7 IDENTITAS", "Profil Usaha",
              "Diisi satu kali oleh admin saat file dibuat.", col=2, span=2)
    p = profil or {}
    section(ws, 6, 2, 2, "DATA USAHA")
    kv(ws, 7, "Nama Usaha", p.get("nama", ""))
    kv(ws, 8, "Nama Pemilik", p.get("pemilik", ""))
    kv(ws, 9, "Jenis Usaha", p.get("jenis", ""))
    kv(ws, 10, "No. HP / WhatsApp", p.get("hp", ""))
    kv(ws, 11, "Alamat", p.get("alamat", ""))
    section(ws, 13, 2, 2, "PEMBUKUAN")
    kv(ws, 14, "Modal Awal / Saldo Kas Awal (Rp)", p.get("modal", 0), RP)
    kv(ws, 15, "Bulan Mulai Pembukuan", p.get("mulai", ""))
    kv(ws, 16, "Nama Admin / Pembuku", p.get("admin", ""))
    note_line(ws, 18, 2, 2,
              "Saldo Kas di Dashboard = Modal Awal di atas + seluruh transaksi yang tercatat.")
    return ws


def sheet_kategori(wb):
    ws = wb.create_sheet("Kategori")
    ws.sheet_properties.tabColor = GOLD
    no_grid(ws)
    widths(ws, {"A": 2, "B": 24, "C": 24, "D": 18, "E": 18, "F": 12, "G": 26})
    page_head(ws, "KASUMKM \u00b7 DAFTAR PILIHAN", "Kategori & Dropdown",
              "Boleh ditambah atau diubah sesuai jenis usaha. Isi tanpa baris kosong di tengah.",
              col=2, span=6)
    table_head(ws, 6, 2, ["Kategori Pemasukan", "Kategori Pengeluaran", "Metode Bayar",
                          "Status", "Jenis", "Semua Kategori (bantu)"])
    for i, v in enumerate(INCOME_CATS, start=7):
        ws.cell(row=i, column=2, value=v)
    for i, v in enumerate(EXPENSE_CATS, start=7):
        ws.cell(row=i, column=3, value=v)
    for i, v in enumerate(METHODS, start=7):
        ws.cell(row=i, column=4, value=v)
    for i, v in enumerate(STATUSES, start=7):
        ws.cell(row=i, column=5, value=v)
    for i, v in enumerate(["Masuk", "Keluar"], start=7):
        ws.cell(row=i, column=6, value=v)
    for i, v in enumerate(INCOME_CATS + EXPENSE_CATS, start=7):
        ws.cell(row=i, column=7, value=v)
    for row in ws.iter_rows(min_row=7, max_row=42, min_col=2, max_col=7):
        for c in row:
            c.font = Font(name=FONT, size=10, color=INK)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", indent=1)
    note_line(ws, 44, 2, 6,
              "Setelah menambah kategori baru, salin juga ke kolom 'Semua Kategori (bantu)' "
              "agar muncul di dropdown sheet Transaksi.", color=AMBER)
    return ws


def sheet_transaksi(wb, rows_data=None):
    ws = wb.create_sheet("Transaksi")
    ws.sheet_properties.tabColor = BLUE
    no_grid(ws)
    cols = ["Tanggal", "Jenis", "Nominal (Rp)", "Keterangan", "Kategori",
            "Metode Bayar", "Status", "Catatan Admin", "Tautan Nota", "Bulan (bantu)"]
    table_head(ws, 1, 1, cols, height=30)
    for i, w in enumerate([12, 10, 16, 34, 20, 14, 16, 26, 22, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{LAST}"

    for r in range(2, LAST + 1):
        ws.row_dimensions[r].height = 18
        ws.cell(row=r, column=1).number_format = DATEF
        ws.cell(row=r, column=3).number_format = RP
        ws.cell(row=r, column=10, value=f'=IF($A{r}="","",TEXT($A{r},"yyyy-mm"))')
        ws.cell(row=r, column=10).font = Font(name=FONT, size=8, color=FAINT)
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c != 10:
                cell.font = Font(name=FONT, size=10, color=INK)

    dv_jenis = DataValidation(type="list", formula1="Kategori!$F$7:$F$8", allow_blank=True,
                              errorTitle="Pilihan tidak tersedia",
                              error="Pilih Masuk atau Keluar dari daftar.")
    dv_kat = DataValidation(type="list", formula1="Kategori!$G$7:$G$42", allow_blank=True)
    dv_met = DataValidation(type="list", formula1="Kategori!$D$7:$D$14", allow_blank=True)
    dv_sta = DataValidation(type="list", formula1="Kategori!$E$7:$E$12", allow_blank=True)
    dv_nom = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
                            allow_blank=True, errorTitle="Nominal tidak valid",
                            error="Tulis angka saja, tanpa Rp dan tanpa titik.")
    for dv, col in ((dv_jenis, "B"), (dv_nom, "C"), (dv_kat, "E"), (dv_met, "F"), (dv_sta, "G")):
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{LAST}")

    rng = f"A2:J{LAST}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$G2="Perlu Perbaikan"'], fill=PatternFill("solid", fgColor=RED_BG)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['AND($A2<>"",$C2="")'], fill=PatternFill("solid", fgColor="FFE0B2")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['AND($A2<>"",$E2="")'], fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add(f"G2:G{LAST}", FormulaRule(
        formula=['$G2="Disetujui"'], font=Font(name=FONT, size=10, bold=True, color=GREEN)))
    ws.conditional_formatting.add(f"B2:B{LAST}", FormulaRule(
        formula=['$B2="Masuk"'], font=Font(name=FONT, size=10, bold=True, color=GREEN)))
    ws.conditional_formatting.add(f"B2:B{LAST}", FormulaRule(
        formula=['$B2="Keluar"'], font=Font(name=FONT, size=10, bold=True, color=RED)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['MOD(ROW(),2)=0'], fill=PatternFill("solid", fgColor=BG)))

    if rows_data:
        for i, row in enumerate(rows_data, start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)
            ws.cell(row=i, column=1).number_format = DATEF
            ws.cell(row=i, column=3).number_format = RP
    return ws


def sheet_data_grafik(wb):
    """Sumber angka untuk kartu KPI & grafik. Disembunyikan dari pemakai biasa."""
    ws = wb.create_sheet("Data Grafik")
    ws.sheet_properties.tabColor = FAINT
    no_grid(ws)
    widths(ws, {"A": 30, "B": 18, "C": 18, "D": 18, "E": 18})
    ws["A1"] = "Data bantu untuk kartu & grafik Dashboard \u2014 tidak perlu diubah"
    ws["A1"].font = Font(name=FONT, size=11, bold=True, color=NAVY)

    masuk = 'SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,{key},Transaksi!$B:$B,"Masuk")'
    keluar = 'SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,{key},Transaksi!$B:$B,"Keluar")'
    saldo = ('IF(\'Profil Usaha\'!$C$14="",0,\'Profil Usaha\'!$C$14)'
             '+SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,"<="&{key},Transaksi!$B:$B,"Masuk")'
             '-SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,"<="&{key},Transaksi!$B:$B,"Keluar")')
    cur, prev = f"Dashboard!{D_KEY}", f"Dashboard!{D_KEY_PREV}"

    table_head(ws, 3, 1, ["Metrik", "Bulan Ini", "Bulan Lalu"], fill=TEAL)
    rows = [
        ("Uang Masuk", f"={masuk.format(key=cur)}", f"={masuk.format(key=prev)}", RP),
        ("Uang Keluar", f"={keluar.format(key=cur)}", f"={keluar.format(key=prev)}", RP),
        ("Laba / Rugi", "=$B$4-$B$5", "=$C$4-$C$5", RP),
        ("Saldo Kas", f"={saldo.format(key=cur)}", f"={saldo.format(key=prev)}", RP),
        ("Jumlah Transaksi", f'=COUNTIFS(Transaksi!$J:$J,{cur})',
         f'=COUNTIFS(Transaksi!$J:$J,{prev})', NUM),
        ("Belum Dikategorikan", f'=COUNTIFS(Transaksi!$J:$J,{cur},Transaksi!$E:$E,"")', None, NUM),
        ("Belum Disetujui", f'=COUNTIFS(Transaksi!$J:$J,{cur},Transaksi!$G:$G,"<>Disetujui")', None, NUM),
        ("Transaksi Terakhir Dicatat", '=IF(COUNT(Transaksi!$A:$A)=0,"-",MAX(Transaksi!$A:$A))', None, DATEF),
    ]
    for i, (label, f_cur, f_prev, fmt) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=r, column=2, value=f_cur).number_format = fmt
        if f_prev:
            ws.cell(row=r, column=3, value=f_prev).number_format = fmt
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BORDER

    # 12 bulan pada tahun terpilih
    table_head(ws, 14, 1, ["Bulan", "Kunci", "Uang Masuk", "Uang Keluar", "Laba / Rugi"], fill=TEAL)
    for i in range(12):
        r = 15 + i
        ws.cell(row=r, column=1, value=f'=CHOOSE({i + 1},{MONTHS_SHORT})')
        ws.cell(row=r, column=2, value=f'=TEXT(DATE(Dashboard!{D_YEAR},{i + 1},1),"yyyy-mm")')
        ws.cell(row=r, column=3,
                value='=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,$B{r},Transaksi!$B:$B,"Masuk")'.format(r=r))
        ws.cell(row=r, column=4,
                value='=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,$B{r},Transaksi!$B:$B,"Keluar")'.format(r=r))
        ws.cell(row=r, column=5, value=f"=$C{r}-$D{r}")
        for c in (3, 4, 5):
            ws.cell(row=r, column=c).number_format = RP
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER

    # pengeluaran per kategori bulan terpilih
    table_head(ws, 29, 1, ["Kategori Pengeluaran", "Jumlah", "Nilai Unik (bantu)"], fill=GOLD)
    for i in range(len(EXPENSE_CATS)):
        r = 30 + i
        ws.cell(row=r, column=1, value=f"=Kategori!$C${7 + i}")
        ws.cell(row=r, column=2,
                value=f'=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,Dashboard!{D_KEY},'
                      f'Transaksi!$B:$B,"Keluar",Transaksi!$E:$E,$A{r})')
        ws.cell(row=r, column=2).number_format = RP
        ws.cell(row=r, column=3, value=f"=IF($B{r}<=0,0,$B{r}+ROW()/1000000)")
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BORDER
    cat_first, cat_last = 30, 29 + len(EXPENSE_CATS)

    # 5 kategori terbesar + Lainnya (sumber grafik donat & daftar di Dashboard)
    table_head(ws, 44, 1, ["Top Kategori Pengeluaran", "Jumlah"], fill=GOLD)
    for k in range(1, 6):
        r = 44 + k
        big = f"LARGE($C${cat_first}:$C${cat_last},{k})"
        ws.cell(row=r, column=1,
                value=f'=IF({big}<=0,"",INDEX($A${cat_first}:$A${cat_last},'
                      f'MATCH({big},$C${cat_first}:$C${cat_last},0)))')
        ws.cell(row=r, column=2,
                value=f'=IF($A{r}="","",INDEX($B${cat_first}:$B${cat_last},'
                      f'MATCH({big},$C${cat_first}:$C${cat_last},0)))')
        ws.cell(row=r, column=2).number_format = RP
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=50, column=1,
            value='=IF(MAX(0,$B$5-SUM($B$45:$B$49))=0,"","Lainnya & belum dikategorikan")')
    ws.cell(row=50, column=2,
            value='=MAX(0,$B$5-SUM($B$45:$B$49))')
    ws.cell(row=50, column=2).number_format = RP
    for c in (1, 2):
        ws.cell(row=50, column=c).border = BORDER

    ws.protection.enable()
    return ws


def sheet_dashboard(wb, ref_month=None):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = GREEN
    no_grid(ws)
    widths(ws, {"A": 2, "B": 23, "C": 15, "D": 2, "E": 23, "F": 15, "G": 3, "H": 2})
    page_head(ws, "KASUMKM \u00b7 DASHBOARD KEUANGAN", "Dashboard",
              "Ubah Bulan & Tahun \u2014 semua kartu, indikator, dan grafik ikut berubah.",
              col=2, span=5,
              title_formula='=IF(\'Profil Usaha\'!$C$7="","Dashboard Keuangan",\'Profil Usaha\'!$C$7)')

    today = ref_month or date.today()
    month_picker(ws, today.month, today.year, "")

    # cermin angka Data Grafik ke baris 8 (disembunyikan) supaya rumus pewarnaan
    # otomatis tetap jalan di Excel maupun Google Sheets (tidak boleh lintas sheet)
    mirror = [("H", "$B$4"), ("I", "$C$4"), ("J", "$B$5"), ("K", "$C$5"),
              ("L", "$B$6"), ("M", "$C$6"), ("N", "$B$7"), ("O", "$C$7")]
    for col, ref in mirror:
        cell = ws[f"{col}8"]
        cell.value = f"={DG}!{ref}"
        cell.number_format = RP
        cell.font = Font(name=FONT, size=8, color=FAINT)

    ws["E7"] = ('=IF($L$8>0,"\u25cf SEHAT \u00b7 laba bulan ini positif",'
                'IF($L$8=0,"\u25cf IMBANG \u00b7 belum ada laba",'
                '"\u25cf WASPADA \u00b7 pengeluaran lebih besar"))')
    ws["E7"].font = Font(name=FONT, size=10, bold=True, color=SLATE)
    ws["E7"].alignment = Alignment(horizontal="right", vertical="center")
    for cond, color, bg in (("$L$8>0", GREEN, GREEN_BG),
                            ("$L$8=0", AMBER, AMBER_BG),
                            ("$L$8<0", RED, RED_BG)):
        ws.conditional_formatting.add("E7:F7", FormulaRule(
            formula=[cond], font=Font(name=FONT, size=10, bold=True, color=color),
            fill=PatternFill("solid", fgColor=bg), stopIfTrue=True))

    card(ws, 10, 2, "UANG MASUK BULAN INI", "=$H$8",
         "$H$8", "$I$8", accent=GREEN, bg=GREEN_BG, good_when_up=True)
    card(ws, 10, 5, "UANG KELUAR BULAN INI", "=$J$8",
         "$J$8", "$K$8", accent=RED, bg=RED_BG, good_when_up=False)
    card(ws, 14, 2, "LABA / RUGI BULAN INI", "=$L$8",
         "$L$8", "$M$8", accent=BLUE, bg=BLUE_BG, good_when_up=True)
    card(ws, 14, 5, "SALDO KAS SAMPAI AKHIR BULAN", "=$N$8",
         "$N$8", "$O$8", accent=AMBER, bg=AMBER_BG, good_when_up=True)
    ws.row_dimensions[13].height = 6
    ws.row_dimensions[17].height = 10

    section(ws, 18, 2, 5, "RINGKASAN CEPAT")
    mini(ws, 19, 2, "Transaksi bulan ini", f"={DG}!$B$8")
    mini(ws, 19, 5, "Belum dikategorikan", f"={DG}!$B$9", bg=YELLOW)
    mini(ws, 20, 2, "Belum disetujui", f"={DG}!$B$10")
    mini(ws, 20, 5, "Terakhir dicatat", f"={DG}!$B$11", fmt=DATEF)
    ws.row_dimensions[21].height = 10

    section(ws, 22, 2, 5, "UANG PALING BANYAK KELUAR BULAN INI")
    table_head(ws, 23, 2, ["Kategori", "Jumlah"], fill=GOLD, height=22)
    table_head(ws, 23, 5, ["Grafik"], fill=GOLD, height=22)
    ws.cell(row=23, column=6).fill = PatternFill("solid", fgColor=GOLD)
    ws.cell(row=23, column=6).border = BORDER
    ws.merge_cells("E23:F23")
    for k in range(6):
        r = 24 + k
        src = 45 + k
        ws.cell(row=r, column=2, value=f'=IF({DG}!$A${src}="","\u2014",{DG}!$A${src})')
        ws.cell(row=r, column=2).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1)
        v = ws.cell(row=r, column=3, value=f'=IF({DG}!$A${src}="","",{DG}!$B${src})')
        v.number_format = RP
        v.font = Font(name=FONT, size=10, bold=True, color=INK)
        v.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        bar = ws.cell(row=r, column=5,
                      value=f'=IF(OR($C{r}="",$C{r}=0),"",REPT("{BAR}",'
                            f'MAX(1,ROUND(16*$C{r}/MAX($C$24:$C$29),0))))')
        bar.font = Font(name=FONT, size=10, color=GOLD)
        bar.alignment = Alignment(vertical="center", indent=1)
        block(ws, r, 2, r, 3, bg=WHITE if k % 2 else BG)
        block(ws, r, 5, r, 6, bg=WHITE if k % 2 else BG)
        ws.row_dimensions[r].height = 19
    note_line(ws, 31, 2, 5,
              "Balok \u2588 makin panjang = pengeluaran makin besar. Grafik lengkap ada di sebelah kanan \u2192")
    ws.protection.enable()
    for a in ("C6", "F6"):
        ws[a].protection = ws[a].protection.copy(locked=False)
    return ws


def style_bar_series(ser, color):
    ser.graphicalProperties = GraphicalProperties(solidFill=color)
    ser.graphicalProperties.line = LineProperties(noFill=True)


def style_line_series(ser, color):
    ser.graphicalProperties = GraphicalProperties()
    ser.graphicalProperties.line = LineProperties(solidFill=color, w=22000)
    ser.smooth = False


def add_dashboard_charts(wb):
    ws = wb["Dashboard"]
    dg = wb["Data Grafik"]
    ak = wb["Arus Kas Harian"]

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.gapWidth = 60
    bar.overlap = -12
    bar.title = "Uang Masuk vs Uang Keluar \u2014 12 Bulan"
    bar.height, bar.width = 8.6, 21
    bar.add_data(Reference(dg, min_col=3, max_col=4, min_row=14, max_row=26), titles_from_data=True)
    bar.set_categories(Reference(dg, min_col=1, min_row=15, max_row=26))
    bar.y_axis.numFmt = NUM
    bar.y_axis.title = None
    bar.legend.position = "b"
    style_bar_series(bar.series[0], "16A34A")
    style_bar_series(bar.series[1], "DC2626")
    ws.add_chart(bar, "I2")

    line = LineChart()
    line.title = "Tren Harian Bulan Ini"
    line.height, line.width = 8.6, 21
    line.add_data(Reference(ak, min_col=2, max_col=3, min_row=4, max_row=35), titles_from_data=True)
    line.set_categories(Reference(ak, min_col=8, min_row=5, max_row=35))
    line.y_axis.numFmt = NUM
    line.legend.position = "b"
    style_line_series(line.series[0], "16A34A")
    style_line_series(line.series[1], "DC2626")
    ws.add_chart(line, "I21")

    dn = DoughnutChart(holeSize=58)
    dn.title = "Komposisi Uang Keluar Bulan Ini"
    dn.height, dn.width = 8.6, 21
    dn.add_data(Reference(dg, min_col=2, min_row=44, max_row=50), titles_from_data=True)
    dn.set_categories(Reference(dg, min_col=1, min_row=45, max_row=50))
    dn.dataLabels = DataLabelList()
    dn.dataLabels.showPercent = True
    dn.dataLabels.showVal = False
    dn.dataLabels.showCatName = False
    dn.dataLabels.showSerName = False
    dn.dataLabels.showLegendKey = False
    dn.dataLabels.showBubbleSize = False
    dn.legend.position = "r"
    palette = ["1D4ED8", "0E7C7B", "C8961E", "B91C1C", "6D28D9", "94A3B8"]
    dn.series[0].data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=c))
        for i, c in enumerate(palette)
    ]
    ws.add_chart(dn, "I40")


def sheet_laba_rugi(wb):
    ws = wb.create_sheet("Laba Rugi")
    ws.sheet_properties.tabColor = "6D28D9"
    no_grid(ws)
    widths(ws, {"A": 2, "B": 34, "C": 18, "D": 14})
    page_head(ws, "KASUMKM \u00b7 LAPORAN BULANAN", "Laporan Laba Rugi",
              "Bulan mengikuti pilihan di sheet Dashboard. Siap dicetak / disimpan sebagai PDF.",
              col=2, span=3)
    ws["B6"] = (f'="Periode: "&CHOOSE(Dashboard!{D_MONTH},{MONTHS_FULL})&" "&Dashboard!{D_YEAR}'
                f'&"  \u00b7  "&IF(\'Profil Usaha\'!$C$7="","(Nama Usaha)",\'Profil Usaha\'!$C$7)')
    ws["B6"].font = Font(name=FONT, size=11, bold=True, color=TEAL)
    ws.merge_cells("B6:D6")

    key = f"Dashboard!{D_KEY}"
    r = 8
    table_head(ws, r, 2, ["PEMASUKAN", "Jumlah (Rp)", "% dari Pemasukan"], fill=TEAL)
    r += 1
    inc_start = r
    for i in range(len(INCOME_CATS)):
        ws.cell(row=r, column=2, value=f"=Kategori!$B${7 + i}")
        ws.cell(row=r, column=3,
                value=f'=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,{key},'
                      f'Transaksi!$B:$B,"Masuk",Transaksi!$E:$E,$B{r})')
        r += 1
    ws.cell(row=r, column=2, value="Belum dikategorikan")
    ws.cell(row=r, column=3,
            value=f'=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,{key},'
                  f'Transaksi!$B:$B,"Masuk",Transaksi!$E:$E,"")')
    inc_end = r
    r += 1
    tot_inc = r
    ws.cell(row=r, column=2, value="TOTAL PEMASUKAN")
    ws.cell(row=r, column=3, value=f"=SUM(C{inc_start}:C{inc_end})")

    r += 2
    table_head(ws, r, 2, ["PENGELUARAN", "Jumlah (Rp)", "% dari Pemasukan"], fill=GOLD)
    r += 1
    exp_start = r
    for i in range(len(EXPENSE_CATS)):
        ws.cell(row=r, column=2, value=f"=Kategori!$C${7 + i}")
        ws.cell(row=r, column=3,
                value=f'=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,{key},'
                      f'Transaksi!$B:$B,"Keluar",Transaksi!$E:$E,$B{r})')
        r += 1
    ws.cell(row=r, column=2, value="Belum dikategorikan")
    ws.cell(row=r, column=3,
            value=f'=SUMIFS(Transaksi!$C:$C,Transaksi!$J:$J,{key},'
                  f'Transaksi!$B:$B,"Keluar",Transaksi!$E:$E,"")')
    exp_end = r
    r += 1
    tot_exp = r
    ws.cell(row=r, column=2, value="TOTAL PENGELUARAN")
    ws.cell(row=r, column=3, value=f"=SUM(C{exp_start}:C{exp_end})")

    for rr in list(range(inc_start, inc_end + 1)) + list(range(exp_start, exp_end + 1)):
        ws.cell(row=rr, column=2).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=rr, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=rr, column=3).number_format = RP
        ws.cell(row=rr, column=3).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=rr, column=4, value=f"=IF($C${tot_inc}=0,\"\",$C{rr}/$C${tot_inc})")
        ws.cell(row=rr, column=4).number_format = '0.0%'
        ws.cell(row=rr, column=4).font = Font(name=FONT, size=9, color=SLATE)
        block(ws, rr, 2, rr, 4, bg=WHITE if rr % 2 else BG)
        ws.row_dimensions[rr].height = 18

    for rr, tint in ((tot_inc, GREEN_BG), (tot_exp, RED_BG)):
        ws.cell(row=rr, column=2).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws.cell(row=rr, column=2).alignment = Alignment(vertical="center", indent=1)
        ws.cell(row=rr, column=3).number_format = RP
        ws.cell(row=rr, column=3).font = Font(name=FONT, size=11, bold=True, color=INK)
        ws.cell(row=rr, column=4, value=f'=IF($C${tot_inc}=0,"",$C{rr}/$C${tot_inc})')
        ws.cell(row=rr, column=4).number_format = '0.0%'
        block(ws, rr, 2, rr, 4, bg=tint)
        ws.row_dimensions[rr].height = 22

    r += 2
    laba_row = r
    block(ws, r, 2, r, 4, bg=NAVY)
    ws.cell(row=r, column=2, value="LABA / RUGI BERSIH").font = Font(
        name=FONT, size=12, bold=True, color=WHITE)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1)
    c = ws.cell(row=r, column=3, value=f"=C{tot_inc}-C{tot_exp}")
    c.number_format = RP
    c.font = Font(name=FONT, size=13, bold=True, color=WHITE)
    ws.cell(row=r, column=4, value=f'=IF($C${tot_inc}=0,"",$C{laba_row}/$C${tot_inc})')
    ws.cell(row=r, column=4).number_format = '0.0%'
    ws.cell(row=r, column=4).font = Font(name=FONT, size=10, bold=True, color=WHITE)
    ws.row_dimensions[r].height = 30

    r += 2
    ws.cell(row=r, column=2, value="Catatan Admin untuk Pemilik Usaha").font = Font(
        name=FONT, size=9, bold=True, color=SLATE)
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=4)
    block(ws, r, 2, r + 2, 4, bg=BG)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    r += 3
    ws.print_area = f"B1:D{r}"
    ws.page_setup.orientation = "portrait"
    ws.protection.enable()
    for rr in range(r - 3, r):
        for cc in range(2, 5):
            cell = ws.cell(row=rr, column=cc)
            cell.protection = cell.protection.copy(locked=False)
    return ws


def sheet_arus_kas(wb):
    ws = wb.create_sheet("Arus Kas Harian")
    ws.sheet_properties.tabColor = "0369A1"
    no_grid(ws)
    widths(ws, {"A": 13, "B": 16, "C": 16, "D": 16, "E": 18, "F": 16, "G": 16, "H": 9})
    ws["A1"] = "Arus Kas Harian"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
    ws["A2"] = ('="Mengikuti Bulan & Tahun di sheet Dashboard \u00b7 "'
                f'&CHOOSE(Dashboard!{D_MONTH},{MONTHS_FULL})&" "&Dashboard!{D_YEAR}')
    ws["A2"].font = Font(name=FONT, size=9, color=SLATE)
    table_head(ws, 4, 1, ["Tanggal", "Uang Masuk", "Uang Keluar", "Selisih Hari Ini",
                          "Saldo Berjalan", "Grafik Masuk", "Grafik Keluar", "Hari (bantu)"])
    for i in range(31):
        r = 5 + i
        ws.cell(row=r, column=1,
                value=f'=IF({i + 1}>DAY(EOMONTH(DATE(Dashboard!{D_YEAR},Dashboard!{D_MONTH},1),0)),"",'
                      f'DATE(Dashboard!{D_YEAR},Dashboard!{D_MONTH},{i + 1}))')
        ws.cell(row=r, column=1).number_format = DATEF
        ws.cell(row=r, column=2,
                value=f'=IF($A{r}="","",SUMIFS(Transaksi!$C:$C,Transaksi!$A:$A,$A{r},'
                      f'Transaksi!$B:$B,"Masuk"))')
        ws.cell(row=r, column=3,
                value=f'=IF($A{r}="","",SUMIFS(Transaksi!$C:$C,Transaksi!$A:$A,$A{r},'
                      f'Transaksi!$B:$B,"Keluar"))')
        ws.cell(row=r, column=4, value=f'=IF($A{r}="","",$B{r}-$C{r})')
        if i == 0:
            saldo = ('=IF($A5="","",IF(\'Profil Usaha\'!$C$14="",0,\'Profil Usaha\'!$C$14)'
                     '+SUMIFS(Transaksi!$C:$C,Transaksi!$A:$A,"<"&$A5,Transaksi!$B:$B,"Masuk")'
                     '-SUMIFS(Transaksi!$C:$C,Transaksi!$A:$A,"<"&$A5,Transaksi!$B:$B,"Keluar")+$D5)')
        else:
            saldo = f'=IF($A{r}="","",$E{r - 1}+$D{r})'
        ws.cell(row=r, column=5, value=saldo)
        ws.cell(row=r, column=6,
                value=f'=IF(OR($A{r}="",$B{r}=0),"",REPT("{BAR}",'
                      f'MAX(1,ROUND(12*$B{r}/MAX($B$5:$B$35),0))))')
        ws.cell(row=r, column=6).font = Font(name=FONT, size=9, color=GREEN)
        ws.cell(row=r, column=7,
                value=f'=IF(OR($A{r}="",$C{r}=0),"",REPT("{BAR}",'
                      f'MAX(1,ROUND(12*$C{r}/MAX($C$5:$C$35),0))))')
        ws.cell(row=r, column=7).font = Font(name=FONT, size=9, color=RED)
        ws.cell(row=r, column=8, value=f'=IF($A{r}="","",DAY($A{r}))')
        ws.cell(row=r, column=8).font = Font(name=FONT, size=8, color=FAINT)
        for col in range(2, 6):
            ws.cell(row=r, column=col).number_format = RP
            ws.cell(row=r, column=col).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=r, column=1).font = Font(name=FONT, size=10, color=INK)
        block(ws, r, 1, r, 8, bg=WHITE if i % 2 else BG)
        ws.row_dimensions[r].height = 18
    ws.conditional_formatting.add("D5:D35", FormulaRule(
        formula=['AND($A5<>"",$D5<0)'], font=Font(name=FONT, size=10, bold=True, color=RED)))
    ws.conditional_formatting.add("E5:E35", FormulaRule(
        formula=['AND($A5<>"",$E5<0)'], font=Font(name=FONT, size=10, bold=True, color=RED),
        fill=PatternFill("solid", fgColor=RED_BG)))

    r = 37
    ws.cell(row=r, column=1, value="TOTAL BULAN").font = Font(name=FONT, size=10, bold=True, color=NAVY)
    for col in (2, 3, 4):
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"=SUM({cl}5:{cl}35)")
        cell.number_format = RP
        cell.font = Font(name=FONT, size=11, bold=True, color=INK)
    block(ws, r, 1, r, 5, bg=BLUE_BG)
    ws.row_dimensions[r].height = 22
    note_line(ws, 39, 1, 5,
              "Baris merah pada Saldo Berjalan = kas minus pada hari itu. Kolom Grafik memakai balok "
              "\u2588 supaya tetap terbaca di HP.")
    ws.protection.enable()
    return ws


def sheet_cek_data(wb):
    ws = wb.create_sheet("Cek Data")
    ws.sheet_properties.tabColor = RED
    no_grid(ws)
    widths(ws, {"A": 2, "B": 20, "C": 13, "D": 16, "E": 32, "F": 56})
    page_head(ws, "KASUMKM \u00b7 KUALITAS DATA", "Cek Data",
              "Pakai Filter pada kolom Masalah \u2192 hilangkan centang (Kosong) untuk melihat "
              "hanya baris yang perlu diperbaiki.", col=2, span=5)
    section(ws, 6, 2, 2, "RINGKASAN")
    mini(ws, 7, 2, "Baris tanpa kategori",
         '=COUNTIFS(Transaksi!$A$2:$A$501,"<>",Transaksi!$E$2:$E$501,"")', bg=YELLOW)
    mini(ws, 8, 2, "Baris tanpa nominal",
         '=COUNTIFS(Transaksi!$A$2:$A$501,"<>",Transaksi!$C$2:$C$501,"")', bg=AMBER_BG)
    mini(ws, 9, 2, "Baris belum disetujui",
         '=COUNTIFS(Transaksi!$A$2:$A$501,"<>",Transaksi!$G$2:$G$501,"<>Disetujui")')
    mini(ws, 10, 2, "Total baris terisi", '=COUNTA(Transaksi!$A$2:$A$501)')

    table_head(ws, 12, 2, ["Baris di Transaksi", "Tanggal", "Nominal", "Keterangan", "Masalah"],
               fill=RED)
    ws.auto_filter.ref = f"B12:F{12 + ROWS}"
    ws.freeze_panes = "B13"
    for i in range(ROWS):
        r = 13 + i
        tr = 2 + i
        ws.cell(row=r, column=2, value=f'=IF(Transaksi!$A{tr}="","",{tr})')
        ws.cell(row=r, column=3, value=f'=IF(Transaksi!$A{tr}="","",Transaksi!$A{tr})')
        ws.cell(row=r, column=3).number_format = DATEF
        ws.cell(row=r, column=4, value=f'=IF(Transaksi!$A{tr}="","",Transaksi!$C{tr})')
        ws.cell(row=r, column=4).number_format = RP
        ws.cell(row=r, column=5, value=f'=IF(Transaksi!$A{tr}="","",Transaksi!$D{tr})')
        ws.cell(row=r, column=6, value=(
            f'=IF(Transaksi!$A{tr}="","",TRIM('
            f'IF(Transaksi!$C{tr}="","Nominal kosong. ","")&'
            f'IF(N(Transaksi!$C{tr})<0,"Nominal negatif. ","")&'
            f'IF(Transaksi!$B{tr}="","Jenis belum diisi. ","")&'
            f'IF(Transaksi!$E{tr}="","Kategori belum diisi. ","")&'
            f'IF(Transaksi!$A{tr}>TODAY(),"Tanggal di masa depan. ","")&'
            f'IF(COUNTIFS(Transaksi!$A$2:$A$501,Transaksi!$A{tr},Transaksi!$C$2:$C$501,Transaksi!$C{tr},'
            f'Transaksi!$D$2:$D$501,Transaksi!$D{tr})>1,"Kemungkinan duplikat. ","")))'))
        for col in range(2, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = Font(name=FONT, size=10, color=INK)
        ws.row_dimensions[r].height = 18
    ws.conditional_formatting.add(f"B13:F{12 + ROWS}", FormulaRule(
        formula=['AND($F13<>"",$B13<>"")'], fill=PatternFill("solid", fgColor=RED_BG)))
    ws.protection.enable()
    return ws


def sheet_contoh(wb, rows):
    ws = wb.create_sheet("Contoh Pengisian")
    ws.sheet_properties.tabColor = FAINT
    no_grid(ws)
    widths(ws, {"A": 2, "B": 13, "C": 10, "D": 16, "E": 34, "F": 22, "G": 16})
    page_head(ws, "KASUMKM \u00b7 CONTOH", "Contoh Pengisian",
              "Hanya contoh. Sheet ini boleh dihapus.", col=2, span=6)
    table_head(ws, 6, 2, ["Tanggal", "Jenis", "Nominal (Rp)", "Keterangan",
                          "Kategori (oleh admin)", "Status"], fill=SLATE)
    for i, row in enumerate(rows, start=7):
        for j, v in enumerate(row, start=2):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=2).number_format = DATEF
        ws.cell(row=i, column=4).number_format = RP
        for j in range(2, 8):
            ws.cell(row=i, column=j).font = Font(name=FONT, size=10, color=INK)
        block(ws, i, 2, i, 7, bg=WHITE if i % 2 else BG)
        ws.row_dimensions[i].height = 18
    note_line(ws, i + 2, 2, 6,
              "Perhatikan: kolom Kategori dan Status boleh kosong saat pemilik usaha mencatat. "
              "Admin yang melengkapi kemudian.")
    return ws


def build_umkm_file(path, profil=None, transaksi=None, ref_month=None):
    wb = Workbook()
    wb.remove(wb.active)
    sheet_petunjuk(wb)
    sheet_profil(wb, profil)
    sheet_transaksi(wb, transaksi)
    sheet_data_grafik(wb)
    sheet_dashboard(wb, ref_month)
    sheet_laba_rugi(wb)
    sheet_arus_kas(wb)
    sheet_kategori(wb)
    sheet_cek_data(wb)
    sheet_contoh(wb, CONTOH_ROWS)
    add_dashboard_charts(wb)
    order = ["Petunjuk", "Profil Usaha", "Transaksi", "Dashboard", "Laba Rugi",
             "Arus Kas Harian", "Kategori", "Cek Data", "Contoh Pengisian", "Data Grafik"]
    wb._sheets = [wb[n] for n in order]
    wb.active = wb.sheetnames.index("Dashboard")
    for name in wb.sheetnames:
        fit_print(wb[name], landscape=name in ("Transaksi", "Cek Data", "Dashboard", "Arus Kas Harian"))
    wb.calculation.fullCalcOnLoad = True
    wb.properties.creator = "KasUMKM"
    wb.properties.title = "Pembukuan UMKM"
    wb.save(path)


# ---------------------------------------------------------------- rekap admin
R_FIRST = 20          # baris pertama data UMKM di sheet Rekap
R_LAST = 39
R_TOTAL = 41


def build_admin_file(path, businesses):
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Petunjuk
    ws = wb.create_sheet("Petunjuk")
    ws.sheet_properties.tabColor = NAVY
    no_grid(ws)
    widths(ws, {"A": 2, "B": 4, "C": 96})
    page_head(ws, "KASUMKM \u00b7 REKAP ADMIN", "Pantau Semua UMKM dari Satu File",
              "Tempel data dari file tiap UMKM, rekap dan tagihan jasa terhitung otomatis.",
              col=2, span=2)
    groups = [
        ("ALUR KERJA MINGGUAN", TEAL, [
            "Buka file Pembukuan-<NamaUsaha>.xlsx \u2192 sheet Transaksi.",
            "Blok data dari kolom Tanggal sampai Status (tanpa baris judul), lalu Copy.",
            "Buka sheet SALIN DATA di file ini, tempel (Paste) di baris kosong paling bawah kolom B.",
            "Tulis nama usaha di kolom A untuk baris-baris yang baru ditempel.",
            "Sheet REKAP otomatis membandingkan seluruh UMKM lewat kartu, indikator, dan grafik.",
            "Sheet TAGIHAN JASA otomatis menghitung biaya jasa pembukuan bulan itu.",
        ]),
        ("CARA BACA REKAP", GREEN, [
            "Kartu di atas: total uang masuk, uang keluar, laba gabungan, dan jumlah UMKM "
            "yang perlu diingatkan \u2014 lengkap dengan \u25b2\u25bc vs bulan lalu.",
            "Kolom Status: \u25cf hijau aktif & rapi, \u25cf kuning masih ada yang belum dikategorikan, "
            "\u25cf merah sudah 3 hari lebih tidak mencatat.",
            "Kolom Grafik Laba memakai balok \u2588 supaya perbandingan tetap terbaca di HP.",
            "Grafik di sebelah kanan: perbandingan uang masuk/keluar dan kontribusi tiap UMKM.",
        ]),
        ("PENTING", AMBER, [
            "Nama usaha di sheet Rekap harus ditulis SAMA PERSIS dengan kolom A sheet Salin Data.",
            "Kolom I (Bulan) di sheet Salin Data adalah kolom bantu rumus \u2014 biarkan saja.",
            "Kapasitas 2.000 baris gabungan. Lakukan rekap setiap akhir minggu supaya tidak menumpuk.",
        ]),
    ]
    r = 6
    for head, color, items in groups:
        section(ws, r, 2, 2, head)
        ws[f"B{r}"].font = Font(name=FONT, size=9, bold=True, color=color)
        r += 1
        for i, text in enumerate(items, start=1):
            ws.cell(row=r, column=2, value=f"{i}.").font = Font(name=FONT, size=9, bold=True, color=color)
            c = ws.cell(row=r, column=3, value=text)
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 26 if len(text) > 95 else 18
            r += 1
        r += 1
    ws.column_dimensions["C"].width = 96

    # ---------------- Salin Data
    ds = wb.create_sheet("Salin Data")
    ds.sheet_properties.tabColor = BLUE
    no_grid(ds)
    table_head(ds, 1, 1, ["Nama Usaha", "Tanggal", "Jenis", "Nominal (Rp)", "Keterangan",
                          "Kategori", "Metode Bayar", "Status", "Bulan (bantu)"], height=30)
    ds.freeze_panes = "A2"
    ds.auto_filter.ref = "A1:I2001"
    for r in range(2, 2002):
        ds.cell(row=r, column=2).number_format = DATEF
        ds.cell(row=r, column=4).number_format = RP
        ds.cell(row=r, column=9, value=f'=IF($B{r}="","",TEXT($B{r},"yyyy-mm"))')
        ds.cell(row=r, column=9).font = Font(name=FONT, size=8, color=FAINT)
    dv = DataValidation(type="list", formula1='"Masuk,Keluar"', allow_blank=True)
    ds.add_data_validation(dv)
    dv.add("C2:C2001")
    ds.conditional_formatting.add("A2:I2001", FormulaRule(
        formula=['AND($B2<>"",$A2="")'], fill=PatternFill("solid", fgColor=RED_BG)))
    ds.conditional_formatting.add("A2:I2001", FormulaRule(
        formula=['AND($B2<>"",$F2="")'], fill=PatternFill("solid", fgColor=YELLOW)))
    for col, w in zip("ABCDEFGHI", (22, 12, 10, 16, 32, 20, 14, 16, 13)):
        ds.column_dimensions[col].width = w

    # ---------------- Rekap
    rk = wb.create_sheet("Rekap")
    rk.sheet_properties.tabColor = GREEN
    no_grid(rk)
    widths(rk, {"A": 24, "B": 17, "C": 17, "D": 17, "E": 14, "F": 17, "G": 15,
                "H": 15, "I": 14, "J": 28, "K": 18, "L": 3})
    page_head(rk, "KASUMKM \u00b7 REKAP ADMIN", "Rekap Semua UMKM",
              "Ubah Bulan & Tahun \u2014 kartu, status, dan grafik ikut berubah.", col=1, span=6)
    # kendali bulan memakai layout yang sama seperti Dashboard UMKM (C6 / F6, kunci di baris 8)
    month_picker(rk, date.today().month, date.today().year, "")
    rk["E7"] = (f'="Jumlah UMKM aktif bulan ini: "&COUNTIFS($E${R_FIRST}:$E${R_LAST},">0")')
    rk["E7"].font = Font(name=FONT, size=10, bold=True, color=TEAL)
    rk["E7"].alignment = Alignment(horizontal="right", vertical="center")

    sd_sum = ('SUMIFS(\'Salin Data\'!$D:$D,\'Salin Data\'!$I:$I,{key},'
              '\'Salin Data\'!$C:$C,"{jenis}")')
    rk["H8"] = f'={sd_sum.format(key="$F$8", jenis="Masuk")}'
    rk["I8"] = f'={sd_sum.format(key="$F$8", jenis="Keluar")}'
    rk["J8"] = "=$H$8-$I$8"
    for a in ("H8", "I8", "J8"):
        rk[a].number_format = RP
        rk[a].font = Font(name=FONT, size=8, color=FAINT)

    card(rk, 10, 2, "TOTAL UANG MASUK", f'={sd_sum.format(key="$C$8", jenis="Masuk")}',
         "$B$11", "$H$8", accent=GREEN, bg=GREEN_BG, good_when_up=True)
    card(rk, 10, 5, "TOTAL UANG KELUAR", f'={sd_sum.format(key="$C$8", jenis="Keluar")}',
         "$E$11", "$I$8", accent=RED, bg=RED_BG, good_when_up=False)
    card(rk, 14, 2, "LABA GABUNGAN SEMUA UMKM", "=$B$11-$E$11",
         "$B$15", "$J$8", accent=BLUE, bg=BLUE_BG, good_when_up=True)
    card(rk, 14, 5, "UMKM PERLU DIINGATKAN",
         f'=COUNTIFS($I${R_FIRST}:$I${R_LAST},">=3")', fmt=NUM,
         accent=AMBER, bg=AMBER_BG, note="belum mencatat 3 hari atau lebih")
    rk.row_dimensions[13].height = 6
    rk.row_dimensions[17].height = 10

    section(rk, 18, 1, 11, "DETAIL PER UMKM")
    table_head(rk, 19, 1, ["Nama Usaha", "Uang Masuk", "Uang Keluar", "Laba / Rugi",
                           "Jumlah Transaksi", "Belum Dikategorikan", "Belum Disetujui",
                           "Transaksi Terakhir", "Hari Tidak Mencatat", "Status",
                           "Grafik Laba"], fill=TEAL, height=32)
    for i in range(R_LAST - R_FIRST + 1):
        r = R_FIRST + i
        rk.cell(row=r, column=1, value=businesses[i] if i < len(businesses) else None)
        rk.cell(row=r, column=2,
                value=f'=IF($A{r}="","",SUMIFS(\'Salin Data\'!$D:$D,\'Salin Data\'!$A:$A,$A{r},'
                      f'\'Salin Data\'!$I:$I,$C$8,\'Salin Data\'!$C:$C,"Masuk"))')
        rk.cell(row=r, column=3,
                value=f'=IF($A{r}="","",SUMIFS(\'Salin Data\'!$D:$D,\'Salin Data\'!$A:$A,$A{r},'
                      f'\'Salin Data\'!$I:$I,$C$8,\'Salin Data\'!$C:$C,"Keluar"))')
        rk.cell(row=r, column=4, value=f'=IF($A{r}="","",$B{r}-$C{r})')
        rk.cell(row=r, column=5,
                value=f'=IF($A{r}="","",COUNTIFS(\'Salin Data\'!$A:$A,$A{r},\'Salin Data\'!$I:$I,$C$8))')
        rk.cell(row=r, column=6,
                value=f'=IF($A{r}="","",COUNTIFS(\'Salin Data\'!$A:$A,$A{r},\'Salin Data\'!$I:$I,$C$8,'
                      f'\'Salin Data\'!$F:$F,""))')
        rk.cell(row=r, column=7,
                value=f'=IF($A{r}="","",COUNTIFS(\'Salin Data\'!$A:$A,$A{r},\'Salin Data\'!$I:$I,$C$8,'
                      f'\'Salin Data\'!$H:$H,"<>Disetujui"))')
        rk.cell(row=r, column=8,
                value=f'=IF($A{r}="","",IF(COUNTIF(\'Salin Data\'!$A:$A,$A{r})=0,"-",'
                      f'SUMPRODUCT(MAX((\'Salin Data\'!$A$2:$A$2001=$A{r})*\'Salin Data\'!$B$2:$B$2001))))')
        rk.cell(row=r, column=9,
                value=f'=IF(OR($A{r}="",NOT(ISNUMBER($H{r}))),"",TODAY()-$H{r})')
        rk.cell(row=r, column=10, value=(
            f'=IF($A{r}="","",IF(NOT(ISNUMBER($H{r})),"\u25cf Belum ada data",'
            f'IF($I{r}>=3,"\u25cf Perlu diingatkan ("&$I{r}&" hari)",'
            f'IF($F{r}>0,"\u25cf Ada yang belum dikategorikan","\u25cf Aktif & rapi"))))'))
        rk.cell(row=r, column=11, value=(
            f'=IF(OR($A{r}="",$D{r}=0),"",REPT("{BAR}",'
            f'MAX(1,ROUND(12*ABS($D{r})/MAX(ABS($D${R_FIRST}:$D${R_LAST})),0))))'))
        for col in (2, 3, 4):
            rk.cell(row=r, column=col).number_format = RP
        rk.cell(row=r, column=8).number_format = DATEF
        rk.cell(row=r, column=9).number_format = '0'
        for col in range(1, 12):
            cell = rk.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = BORDER
        rk.cell(row=r, column=1).alignment = Alignment(vertical="center", indent=1)
        rk.cell(row=r, column=10).font = Font(name=FONT, size=9, bold=True, color=SLATE)
        rk.cell(row=r, column=11).font = Font(name=FONT, size=10, color=BLUE)
        rk.row_dimensions[r].height = 19

    rk.cell(row=R_TOTAL, column=1, value="TOTAL SEMUA UMKM").font = Font(
        name=FONT, size=10, bold=True, color=NAVY)
    for col in range(2, 8):
        cl = get_column_letter(col)
        cell = rk.cell(row=R_TOTAL, column=col, value=f"=SUM({cl}{R_FIRST}:{cl}{R_LAST})")
        cell.font = Font(name=FONT, size=11, bold=True, color=INK)
        cell.number_format = RP if col <= 4 else NUM
    block(rk, R_TOTAL, 1, R_TOTAL, 11, bg=BLUE_BG)
    rk.row_dimensions[R_TOTAL].height = 22

    dat = f"A{R_FIRST}:K{R_LAST}"
    rk.conditional_formatting.add(dat, FormulaRule(
        formula=[f'AND($I{R_FIRST}<>"",$I{R_FIRST}>=3)'], fill=PatternFill("solid", fgColor=RED_BG)))
    rk.conditional_formatting.add(dat, FormulaRule(
        formula=[f'AND($F{R_FIRST}<>"",$F{R_FIRST}>0)'], fill=PatternFill("solid", fgColor=YELLOW)))
    rk.conditional_formatting.add(f"J{R_FIRST}:J{R_LAST}", FormulaRule(
        formula=[f'AND($I{R_FIRST}<>"",$I{R_FIRST}>=3)'],
        font=Font(name=FONT, size=9, bold=True, color=RED), stopIfTrue=True))
    rk.conditional_formatting.add(f"J{R_FIRST}:J{R_LAST}", FormulaRule(
        formula=[f'AND($A{R_FIRST}<>"",$F{R_FIRST}>0)'],
        font=Font(name=FONT, size=9, bold=True, color=AMBER), stopIfTrue=True))
    rk.conditional_formatting.add(f"J{R_FIRST}:J{R_LAST}", FormulaRule(
        formula=[f'$J{R_FIRST}="\u25cf Aktif & rapi"'],
        font=Font(name=FONT, size=9, bold=True, color=GREEN), stopIfTrue=True))
    rk.conditional_formatting.add(f"D{R_FIRST}:D{R_LAST}", FormulaRule(
        formula=[f'AND($A{R_FIRST}<>"",$D{R_FIRST}<0)'],
        font=Font(name=FONT, size=10, bold=True, color=RED)))
    note_line(rk, R_TOTAL + 2, 1, 11,
              "Baris merah = 3 hari atau lebih tidak mencatat (perlu diingatkan). "
              "Baris kuning = masih ada transaksi yang belum dikategorikan. "
              "Nama usaha di kolom A harus sama persis dengan sheet Salin Data.")

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.gapWidth = 60
    bar.overlap = -12
    bar.title = "Uang Masuk vs Uang Keluar per UMKM"
    bar.height, bar.width = 9, 22
    bar.add_data(Reference(rk, min_col=2, max_col=3, min_row=19, max_row=R_LAST), titles_from_data=True)
    bar.set_categories(Reference(rk, min_col=1, min_row=R_FIRST, max_row=R_LAST))
    bar.y_axis.numFmt = NUM
    bar.legend.position = "b"
    style_bar_series(bar.series[0], "16A34A")
    style_bar_series(bar.series[1], "DC2626")
    rk.add_chart(bar, "M3")

    laba = BarChart()
    laba.type = "bar"
    laba.grouping = "clustered"
    laba.gapWidth = 50
    laba.title = "Laba / Rugi per UMKM (bulan terpilih)"
    laba.height, laba.width = 9, 22
    laba.add_data(Reference(rk, min_col=4, min_row=19, max_row=R_LAST), titles_from_data=True)
    laba.set_categories(Reference(rk, min_col=1, min_row=R_FIRST, max_row=R_LAST))
    laba.x_axis.numFmt = NUM
    laba.legend = None
    style_bar_series(laba.series[0], "1D4ED8")
    rk.add_chart(laba, "M22")

    # ---------------- Tagihan Jasa
    tg = wb.create_sheet("Tagihan Jasa")
    tg.sheet_properties.tabColor = GOLD
    no_grid(tg)
    widths(tg, {"A": 2, "B": 36, "C": 17, "D": 18, "E": 16, "F": 16, "G": 17, "H": 16})
    page_head(tg, "KASUMKM \u00b7 PENDAPATAN JASA", "Tagihan Jasa Pembukuan",
              "Bulan mengikuti sheet Rekap. Ubah tarif di bawah sesuai kesepakatan.", col=2, span=6)
    kv(tg, 6, "Tarif Dasar per UMKM / bulan (Rp)", 50000, RP)
    kv(tg, 7, "Batas Transaksi Termasuk Tarif Dasar", 100, NUM)
    kv(tg, 8, "Biaya Tambahan per Transaksi Kelebihan (Rp)", 300, RP)
    tg["E6"] = f'=CHOOSE(Rekap!{D_MONTH},{MONTHS_FULL})&" "&Rekap!{D_YEAR}'
    tg["E6"].font = Font(name=FONT, size=12, bold=True, color=TEAL)
    tg.row_dimensions[9].height = 10

    table_head(tg, 10, 2, ["Nama Usaha", "Jumlah Transaksi", "Kelebihan Transaksi", "Tarif Dasar",
                           "Biaya Tambahan", "Total Tagihan", "Status Bayar"], fill=GOLD, height=30)
    first, last = 11, 11 + (R_LAST - R_FIRST)
    for i in range(R_LAST - R_FIRST + 1):
        r = first + i
        src = R_FIRST + i
        tg.cell(row=r, column=2, value=f'=IF(Rekap!$A{src}="","",Rekap!$A{src})')
        tg.cell(row=r, column=3, value=f'=IF($B{r}="","",Rekap!$E{src})')
        tg.cell(row=r, column=4, value=f'=IF($B{r}="","",MAX(0,$C{r}-$C$7))')
        tg.cell(row=r, column=5, value=f'=IF($B{r}="","",IF($C{r}=0,0,$C$6))')
        tg.cell(row=r, column=6, value=f'=IF($B{r}="","",$D{r}*$C$8)')
        tg.cell(row=r, column=7, value=f'=IF($B{r}="","",$E{r}+$F{r})')
        for col in (5, 6, 7):
            tg.cell(row=r, column=col).number_format = RP
        for col in range(2, 9):
            cell = tg.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = BORDER
        tg.cell(row=r, column=7).font = Font(name=FONT, size=10, bold=True, color=INK)
        tg.row_dimensions[r].height = 19
    dvb = DataValidation(type="list", formula1='"Belum Bayar,Sudah Bayar"', allow_blank=True)
    tg.add_data_validation(dvb)
    dvb.add(f"H{first}:H{last}")
    tg.conditional_formatting.add(f"H{first}:H{last}", FormulaRule(
        formula=[f'$H{first}="Sudah Bayar"'], font=Font(name=FONT, size=10, bold=True, color=GREEN),
        fill=PatternFill("solid", fgColor=GREEN_BG)))
    tg.conditional_formatting.add(f"H{first}:H{last}", FormulaRule(
        formula=[f'$H{first}="Belum Bayar"'], font=Font(name=FONT, size=10, bold=True, color=RED),
        fill=PatternFill("solid", fgColor=RED_BG)))

    tot = last + 2
    card(tg, tot, 2, "TOTAL PENDAPATAN JASA BULAN INI", f"=SUM(G{first}:G{last})",
         accent=GREEN, bg=GREEN_BG, note="dari seluruh UMKM aktif")
    card(tg, tot, 5, "BELUM DIBAYAR",
         f'=SUM(G{first}:G{last})-SUMIFS($G${first}:$G${last},$H${first}:$H${last},"Sudah Bayar")',
         accent=AMBER, bg=AMBER_BG, note="tagihan yang masih perlu ditagih")
    mini(tg, tot + 4, 2, "Sudah dibayar",
         f'=SUMIFS($G${first}:$G${last},$H${first}:$H${last},"Sudah Bayar")', fmt=RP, bg=GREEN_BG)
    mini(tg, tot + 4, 5, "Jumlah UMKM ditagih", f'=COUNTIF($B${first}:$B${last},"?*")', bg=BG)

    for name in wb.sheetnames:
        fit_print(wb[name], landscape=name in ("Salin Data", "Rekap", "Tagihan Jasa"))
    wb.active = wb.sheetnames.index("Rekap")
    wb.calculation.fullCalcOnLoad = True
    wb.properties.creator = "KasUMKM"
    wb.properties.title = "Rekap Admin KasUMKM"
    wb.save(path)


# ---------------------------------------------------------------- data contoh
CONTOH_ROWS = [
    (date(2026, 6, 1), "Masuk", 450000, "penjualan hari ini", "Penjualan Tunai", "Disetujui"),
    (date(2026, 6, 1), "Keluar", 150000, "beli gula 5 kg", "Bahan Baku", "Disetujui"),
    (date(2026, 6, 2), "Masuk", 380000, "jualan", None, None),
    (date(2026, 6, 2), "Keluar", 25000, "bensin antar pesanan", None, None),
    (date(2026, 6, 3), "Masuk", 620000, "pesanan katering 20 box", "Pendapatan Jasa", "Disetujui"),
    (date(2026, 6, 3), "Keluar", 300000, "belanja pasar", "Bahan Baku", "Disetujui"),
    (date(2026, 6, 4), "Keluar", 500000, "bayar listrik", "Listrik & Air", "Disetujui"),
]


def sample_transactions():
    """±2 bulan transaksi contoh untuk Toko Maju."""
    import random
    random.seed(7)
    rows = []
    start = date(2026, 5, 1)
    for d in range(0, 61):
        day = start + timedelta(days=d)
        for _ in range(random.choice([1, 1, 2])):
            amount = random.randrange(150, 900) * 1000
            kat = random.choice(["Penjualan Tunai", "Penjualan Tunai", "Penjualan Online"])
            ket = random.choice(["penjualan harian", "jualan warung", "pesanan online",
                                 "jualan sore", "penjualan tunai kasir"])
            reviewed = day < date(2026, 6, 20)
            rows.append([day, "Masuk", amount, ket,
                         kat if reviewed else None,
                         "Tunai" if reviewed else None,
                         "Disetujui" if reviewed else None, None, None])
        if random.random() < 0.8:
            pilihan = [("beli bahan baku", "Bahan Baku"), ("belanja pasar", "Bahan Baku"),
                       ("bensin motor kirim", "Transport & Kirim"), ("beli plastik kemasan", "Kemasan"),
                       ("pulsa internet", "Pulsa & Internet"), ("beli stok minuman", "Barang Dagang")]
            ket, kat = random.choice(pilihan)
            amount = random.randrange(20, 350) * 1000
            reviewed = day < date(2026, 6, 20)
            rows.append([day, "Keluar", amount, ket,
                         kat if reviewed else None,
                         "Tunai" if reviewed else None,
                         "Disetujui" if reviewed else None, None, None])
        if day.day == 5:
            rows.append([day, "Keluar", 1500000, "gaji karyawan", "Gaji & Upah", "Transfer Bank",
                         "Disetujui", None, None])
            rows.append([day, "Keluar", 1000000, "sewa kios", "Sewa Tempat", "Transfer Bank",
                         "Disetujui", None, None])
        if day.day == 10:
            rows.append([day, "Keluar", 320000, "bayar listrik & air", "Listrik & Air", "Transfer Bank",
                         "Disetujui", None, None])
    rows.sort(key=lambda x: x[0])
    return rows


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    build_umkm_file(f"{OUT_DIR}/Pembukuan-Template.xlsx", profil={"mulai": "", "admin": ""})
    build_umkm_file(
        f"{OUT_DIR}/Pembukuan-TokoMaju-Contoh.xlsx",
        profil={"nama": "Toko Maju", "pemilik": "Bu Siti", "jenis": "Warung / Toko Kelontong",
                "hp": "0812-3456-7890", "alamat": "Jl. Merdeka No. 12", "modal": 5000000,
                "mulai": "Mei 2026", "admin": "Admin KasUMKM"},
        transaksi=sample_transactions(), ref_month=date(2026, 6, 1))
    build_admin_file(f"{OUT_DIR}/Rekap-Admin.xlsx",
                     ["Toko Maju", "Kedai Nusantara", "Laundry Bersih", "Sinar Jaya Online",
                      "Kreatif Digital"])
    for f in sorted(os.listdir(OUT_DIR)):
        print(f, os.path.getsize(os.path.join(OUT_DIR, f)), "bytes")


if __name__ == "__main__":
    main()
