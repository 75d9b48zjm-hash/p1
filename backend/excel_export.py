import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

NAVY = "0F172A"
GREEN = "059669"
RED = "DC2626"
SLATE = "475569"
LIGHT = "F1F5F9"
WHITE = "FFFFFF"
RUPIAH = '"Rp"#,##0'

_thin = Side(style="thin", color="E2E8F0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _brand(ws, text, sub):
    ws["B2"] = "KasUMKM"
    ws["B2"].font = Font(bold=True, size=10, color=GREEN)
    ws["B3"] = text
    ws["B3"].font = Font(bold=True, size=18, color=NAVY)
    ws["B4"] = sub
    ws["B4"].font = Font(size=10, color=SLATE)


def _header_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def build_business_excel(business, report, monthly, transactions, period_label):
    """Excel sederhana: ringkasan angka + grafik + tabel transaksi. Tanpa macro."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in "BCDE":
        ws.column_dimensions[col].width = 20

    _brand(ws, business["name"], f"Laporan periode {period_label}")

    kpis = [
        ("Saldo Awal", report["opening_balance"], SLATE),
        ("Uang Masuk", report["total_income"], GREEN),
        ("Uang Keluar", report["total_expense"], RED),
        ("Laba Bersih", report["net_profit"], GREEN if report["net_profit"] >= 0 else RED),
        ("Saldo Akhir", report["closing_balance"], NAVY),
    ]
    r = 6
    ws[f"B{r}"] = "RINGKASAN KEUANGAN"
    ws[f"B{r}"].font = Font(bold=True, size=11, color=NAVY)
    r += 1
    for label, val, color in kpis:
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = Font(size=10, color=SLATE)
        c = ws[f"C{r}"]
        c.value = val
        c.number_format = RUPIAH
        c.font = Font(bold=True, size=12, color=color)
        r += 1

    mstart = r + 1
    ws[f"B{mstart}"] = "6 BULAN TERAKHIR"
    ws[f"B{mstart}"].font = Font(bold=True, size=11, color=NAVY)
    hdr = mstart + 1
    for i, h in enumerate(["Bulan", "Uang Masuk", "Uang Keluar", "Laba"]):
        _header_cell(ws.cell(row=hdr, column=2 + i), h)
    for j, m in enumerate(monthly):
        rr = hdr + 1 + j
        ws.cell(row=rr, column=2, value=m["month"]).border = BORDER
        for k, key in enumerate(["income", "expense", "profit"]):
            c = ws.cell(row=rr, column=3 + k, value=m[key])
            c.number_format = RUPIAH
            c.border = BORDER
    last = hdr + len(monthly)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Uang Masuk vs Uang Keluar"
    chart.height = 7.5
    chart.width = 17
    data = Reference(ws, min_col=3, max_col=4, min_row=hdr, max_row=last)
    cats = Reference(ws, min_col=2, min_row=hdr + 1, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.numFmt = RUPIAH
    ws.add_chart(chart, f"F{mstart}")

    # ---- Transaksi sheet ----
    ts = wb.create_sheet("Transaksi")
    ts.sheet_view.showGridLines = False
    tcols = [("Tanggal", 14), ("Jenis", 12), ("Kategori", 20), ("Deskripsi", 36), ("Metode", 16), ("Nominal", 16)]
    ts.column_dimensions["A"].width = 2
    ts["B2"] = f"{business['name']} — Daftar Transaksi ({period_label})"
    ts["B2"].font = Font(bold=True, size=12, color=NAVY)
    for i, (h, w) in enumerate(tcols):
        ts.column_dimensions[get_column_letter(2 + i)].width = w
        _header_cell(ts.cell(row=4, column=2 + i), h)
    for j, t in enumerate(transactions):
        rr = 5 + j
        is_income = t["type"] == "income"
        vals = [
            t.get("date", ""),
            "Uang Masuk" if is_income else "Uang Keluar",
            t.get("category", ""),
            t.get("description", ""),
            t.get("payment_method", ""),
            t.get("amount", 0),
        ]
        for k, v in enumerate(vals):
            c = ts.cell(row=rr, column=2 + k, value=v)
            c.border = BORDER
            if k == 5:
                c.number_format = RUPIAH
                c.font = Font(size=10, bold=True, color=GREEN if is_income else RED)
            else:
                c.font = Font(size=10, color=NAVY)
        if j % 2 == 1:
            for k in range(6):
                ts.cell(row=rr, column=2 + k).fill = PatternFill("solid", fgColor=LIGHT)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
